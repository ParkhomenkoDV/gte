import time
import math
import numpy as np
import matplotlib.pyplot as plt
import scipy
import openpyxl

start_time = time.monotonic()  # начало отсчета времени


def print_matrix(TABLE):
    for arr in TABLE:
        for el in arr:
            print(el, end=' ')
        print()


ORIGIN = 1  # начало отсчета начиная c
Error = "Ошибки, баги и предупреждения:"

# --- ДАНО ---

# передаточное отношение
u = 7
# ресурс (ч)
Lh = 40000
# момент (Нм)
T1 = 400
# угол наклона зубьев (рад)
beta = 0
# твердость сердцевины (0) и зубьев (1) (HRC)
H0 = 48
H1 = 52
# квалитет точности
Quality = 8

modul1 = []  # модули 1 ряда
modul2 = []  # модули 2 ряда
HB = []  # твердость по Бринелю
HRC = []  # твердость по Роквеллу
z2 = []
z1 = []
x1 = []

# считка модулей из excel-файла
wb = openpyxl.reader.excel.load_workbook(filename="Модули зубчатого зацепления.xlsx", data_only=True)
wb.active = 0  # первый лист
sheet = wb.active
for i in range(1 + ORIGIN, 35 + ORIGIN): modul1.append(sheet['A' + str(i)].value)
for i in range(1 + ORIGIN, 34 + ORIGIN): modul2.append(sheet['B' + str(i)].value)

# считка твердостей из excel-файла
wb = openpyxl.reader.excel.load_workbook(filename="Твердость.xlsx", data_only=True)
wb.active = 0  # первый лист
sheet = wb.active
for i in range(1 + ORIGIN, 92 + ORIGIN):
    HB.append(sheet['A' + str(i)].value)
    HRC.append(sheet['B' + str(i)].value)


def HRC2HB(HH): return scipy.interpolate.interp1d(HRC, HB, kind='cubic')(HH)


def HB2HRC(HH): return scipy.interpolate.interp1d(HB, HRC, kind='cubic')(HH)


#
wb = openpyxl.reader.excel.load_workbook(filename="Коэф. смещения x1.xlsx", data_only=True)
wb.active = 0  # первый лист
sheet = wb.active
for row in range(2, 11 + ORIGIN):
    z1.append(sheet[row][1 - ORIGIN].value)
for col in range(6 - ORIGIN, 10 + ORIGIN):
    z2.append(sheet[1][col].value)
    '''
for row in range(2, 11 + ORIGIN):
    for col in range(6 - ORIGIN, 10 + ORIGIN):
        x1.append(sheet[row].value)
        '''
# print(z1)
# print(z2)
# print(x1)

# print(x1)

# def x1(zz1, zz2):
# return scipy.interpolate.interp2d(z1, z2, x1, kind='linear')(zz1, zz2)


'''
plt.plot(HB, HRC, color = 'orange', linestyle='-')
plt.show()
'''

if H1 > 50:
    psi = 0.25
elif 37.8 <= H1 <= 50:
    psi = 0.315
else:
    psi = 0.4

if H1 > 52:
    sHlim = min(23 * H1, 1050)
elif 40 <= H1 <= 52:
    sHlim = 17 * H1 + 200
else:
    sHlim = 2 * 37.8 + 70

SH = 1.2

sg = \
    {
        'i': [],
        'z1': [],
        'z2': [],
        'z3': [],
        'u': [],
        'epsilon': [],
        'm': [],
        'V': [],
        'aw': [],
        'd1': [],
        'd2': [],
        'd3': [],
        'b1': [],
        'b2': [],
        'b3': [],
    }


# кинематический расчет
def Kinematic_Calculation():
    global ii
    i = 0
    ii = 0
    for z1 in range(17, 120 + ORIGIN):
        for n2 in range(2, 6 + ORIGIN):  # количество сателитов
            for m in modul1:
                ii = ii + 1
                z3 = round(z1 * (u - 1))
                z2 = round((z3 - z1) / 2)
                # условия сборки и соосности
                if (((z1 + z2) / n2) - round((z1 + z2) / n2) == 0) and (z2 + 2) < (z1 + z2) * math.sin(math.pi / n2):
                    uu = z3 / z1 + 1
                    epsilon = abs(u - uu) / u * 100
                    aw = m * (z1 + z2) / 2
                    d1 = z1 * m / math.cos(beta)
                    d2 = 2 * aw - d1
                    d3 = d1 + 2 * d2
                    d1a = d1 + 2 * m
                    d2a = d2 + 2 * m
                    d1f = d1 - 2.5 * m
                    d2f = d2 - 2.5 * m
                    b3 = round(psi * aw)
                    b2 = b3 + 3
                    b1 = 1.1 * b2
                    V = round(d1 * b1 + n2 * d2 * b2 + d3 * b3)
                    # условие соосности и соседства
                    if (aw * math.sin(math.pi / n2) > d1a / 2):
                        i += 1
                        # TABLE.append([i, z1, z2, z3, uu, epsilon, m, V, aw, d1, d2, d3, b1, b2, b3])
                        sg['i'].append(ii)
                        sg['z1'].append(z1)
                        sg['z2'].append(z2)
                        sg['z3'].append(z3)
                        sg['u'].append(uu)
                        sg['epsilon'].append(epsilon)
                        sg['m'].append(m)
                        sg['V'].append(V)
                        sg['aw'].append(aw)
                        sg['d1'].append(d1)
                        sg['d2'].append(d2)
                        sg['d3'].append(d3)
                        sg['b1'].append(b1)
                        sg['b2'].append(b2)
                        sg['b3'].append(b3)


Kinematic_Calculation()

print()
print("Проверено ", ii, " вариантов")
print("Выбрано ", len(sg['i']), " вариантов")
print()

print(sg['i'])
print(sg['z1'])
print(sg['z2'])
print(sg['z3'])
print(sg['u'])
print(sg['epsilon'])
print(sg['m'])
print(sg['V'])
print(sg['aw'])
print(sg['d1'])
print(sg['d2'])
print(sg['d3'])
print(sg['b1'])
print(sg['b2'])
print(sg['b3'])

# print_matrix(TABLE)

zs = 9600 if beta == 0 else 8400
sH = sHlim / SH

print()
print("Program ended in", round(time.monotonic() - start_time, 4))
print(Error)
