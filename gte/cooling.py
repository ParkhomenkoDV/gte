import math
import os
from math import degrees, radians, sqrt, pi, sin, cos

import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd
from scipy import interpolate
from scipy.optimize import root

from Газодинамические_функции import m as m_const
from Газодинамические_функции import paskal, tau, q
from Характеристики_рабочих_тел import Cp_возд as Cp_возд
from Характеристики_рабочих_тел import R_продуктов_сгорания_природного_газа


# Матрица поворота
def rotation_matrix(angle, x, x_transfer, y, y_transfer):
    A_ = np.array([[cos(angle), sin(angle)], [-sin(angle), cos(angle)]])
    X_ = np.zeros(len(x))
    Y_ = np.zeros(len(x))
    for i_ in range(len(x)):
        B_ = np.array([[(x[i_] - x_transfer)], [(y[i_] - y_transfer)]])
        [X_[i_], Y_[i_]] = np.matmul(A_, B_)
    return X_, Y_


def lambda_ад(s_отн_, lambda_2_ад):
    if (-0.891 <= s_отн_ <= 0.91) and (0.77 <= lambda_2_ад <= 1.39):
        X_s_отн = np.array(
            [-0.891, -0.841, -0.784, -0.729, -0.674, -0.613, -0.563, -0.497, -0.442, -0.390, -0.317, -0.25, -0.163,
             -0.113, -0.061, 0, 0.061, 0.172, 0.262, 0.358, 0.483, 0.619, 0.680, 0.758, 0.91])
        Y_lambda_2_ад = np.array([0.77, 0.86, 0.97, 1.06, 1.23, 1.39])
        Z_lambda_ад = np.array([[0.791, 0.805, 0.783, 0.801, 0.811, 0.816, 0.836, 0.842, 0.826, 0.770, 0.670, 0.581,
                                 0.458, 0.358, 0.235, 0, 0.165, 0.199, 0.194, 0.214, 0.297, 0.430, 0.500, 0.573, 0.717],
                                [0.898, 0.933, 0.930, 0.920, 0.913, 0.909, 0.968, 0.959, 0.909, 0.845, 0.727, 0.616,
                                 0.490, 0.401, 0.235, 0, 0.165, 0.199, 0.216, 0.246, 0.336, 0.469, 0.542, 0.624, 0.784],
                                [0.999, 1.026, 1.007, 1.020, 0.997, 1.011, 1.183, 1.054, 0.936, 0.845, 0.727, 0.616,
                                 0.490, 0.401, 0.235, 0, 0.165, 0.199, 0.216, 0.246, 0.366, 0.506, 0.568, 0.659, 0.820],
                                [1.154, 1.138, 1.113, 1.081, 1.191, 1.327, 1.208, 1.054, 0.936, 0.845, 0.727, 0.619,
                                 0.490, 0.401, 0.235, 0, 0.165, 0.199, 0.216, 0.246, 0.366, 0.506, 0.568, 0.659, 0.820],
                                [1.238, 1.185, 1.243, 1.417, 1.491, 1.349, 1.208, 1.054, 0.936, 0.845, 0.727, 0.619,
                                 0.490, 0.401, 0.235, 0, 0.165, 0.199, 0.216, 0.246, 0.366, 0.506, 0.568, 0.659, 0.820],
                                [1.507, 1.526, 1.486, 1.595, 1.502, 1.349, 1.208, 1.054, 0.936, 0.845, 0.727, 0.619,
                                 0.490, 0.401, 0.235, 0, 0.165, 0.199, 0.216, 0.246, 0.366, 0.506, 0.568, 0.659,
                                 0.820]])
        return interpolate.interp2d(X_s_отн, Y_lambda_2_ад, Z_lambda_ад, kind='linear')(s_отн_, lambda_2_ад)
    else:
        if (s_отн < -0.891 or s_отн > 0.91) and (lambda_2_ад < 0.77 or lambda_2_ад > 1.39):
            return print('Нет данных lambda_ад для s_отн = ' + str(s_отн) + '.\n Нет данных lambda_ад для lambda_2_ад = ' + str(lambda_2_ад) + '.')
        elif s_отн < -0.891 or s_отн > 0.91:
            return print('Нет данных lambda_ад для s_отн = ' + str(s_отн) + '.')
        elif lambda_2_ад < 0.77 or lambda_2_ад > 1.39:
            return print('Нет данных lambda_ад для lambda_2_ад = ' + str(lambda_2_ад) + '.')


# Зависимость коэф. А от угла поворота потока
def коэф_A(epsilon_):
    if radians(25) <= epsilon_ <= radians(60):
        X_epsilon = np.array([25, 30, 35, 40, 45, 50, 55, 60]) * pi / 180
        Y_A = np.array([1.1, 0.989, 0.906, 0.856, 0.823, 0.802, 0.789, 0.788])
        return interpolate.interp1d(X_epsilon, Y_A, kind='cubic')(epsilon_)
    else:
        print('Угол поворота потока epsilon = ' + str(degrees(epsilon_)) + '. epsilon ∈ [25..60]')


def my_flip_matr(matr, n_flip):
    return np.hstack((np.flip(matr[0: n_flip], 0), matr[n_flip:]))


def combining_functions(fun, x_s_профиля, l_спинки, l_корыта, N_спинки, N_профиля, T_охл, x):
    for p in range(len(x_s_профиля)):
        if (p == N_спинки - 1) and (-l_спинки <= x < -x_s_профиля[p]):
            return fun[p](-x)
        elif (0 <= p < N_спинки - 1) and (-x_s_профиля[p + 1] <= x < -x_s_профиля[p]):
            return fun[p](-x)
        elif (N_спинки <= p < N_профиля - 1) and (x_s_профиля[p] < x <= x_s_профиля[p + 1]):
            return fun[p](x)
        elif -x_s_профиля[0] < x < x_s_профиля[N_спинки]:
            return T_охл
        elif (p == N_профиля - 1) and (x_s_профиля[p] < x <= l_корыта):
            return fun[p](x)


def phi_1_fun(Xn_D, Re_D):
    if Re_D <= 3 * 10 ** 3:
        x_Xn_D = np.array([1.5, 2, 3.255, 4.382, 4.569, 4.806, 5.091, 5.492, 6.025, 6.693, 7.681, 9.238, 12.327])
        y_phi_1 = np.array([0.7, 0.4, 0.2, 0.1, 0.09, 0.08, 0.07, 0.06, 0.05, 0.04, 0.03, 0.02, 0.01])
    elif Re_D > 3 * 10 ** 3:
        x_Xn_D = np.array(
            [1.5, 2, 2.775, 3.620, 3.756, 3.922, 4.111, 4.358, 4.655, 4.955, 5.502, 6.299, 8.123, 8.855, 9.868, 10.499,
             11.384,
             12.552])
        y_phi_1 = np.array(
            [0.85, 0.52, 0.2, 0.1, 0.09, 0.08, 0.07, 0.06, 0.05, 0.04, 0.03, 0.02, 0.01, 0.008, 0.006, 0.005, 0.004,
             0.003])
    return interpolate.interp1d(x_Xn_D, y_phi_1, kind='linear')(Xn_D)


def phi_2_fun(Комплекс, Re_D):
    if Re_D <= 3 * 10 ** 3:
        x_Комплекс = np.array([0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0, 1.1, 1.2, 1.3, 1.4])
        y_phi_2 = np.array(
            [0.400, 0.345, 0.307, 0.274, 0.243, 0.219, 0.195, 0.175, 0.157, 0.140, 0.123, 0.105, 0.089, 0.074,
             0.058]) + 0.6
    elif Re_D > 3 * 10 ** 3:
        x_Комплекс = np.array([0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0, 1.1, 1.2, 1.3, 1.4])
        y_phi_2 = np.array(
            [0.500, 0.445, 0.405, 0.368, 0.333, 0.304, 0.276, 0.252, 0.231, 0.210, 0.188, 0.167, 0.148, 0.128,
             0.108]) + 0.5
    return interpolate.interp1d(x_Комплекс, y_phi_2, kind='linear')(Комплекс)


def m_exp(Xn_D, Re_D):
    if Re_D <= 3 * 10 ** 3:
        x_Xn_D = np.array([0, 2, 4, 6, 8, 10, 12])
        y_m_exp = np.array([0, 0.106, 0.180, 0.234, 0.273, 0.303, 0.328]) + 0.5
    elif 3 * 10 ** 3 < Re_D:
        x_Xn_D = np.array([0, 2, 4, 6, 8, 10, 12])
        y_m_exp = np.array([0, 0.133, 0.244, 0.333, 0.395, 0.440, 0.465]) + 0.5

    return interpolate.interp1d(x_Xn_D, y_m_exp)(Xn_D)


def delta_teta_пл_gamma_simple_fun(teta_пл_0, uв_uг, gamma, x):
    if 30 < gamma <= 90:
        if 75 < gamma <= 90:
            X_teta_пл_0 = np.array([0, 0.2, 0.4, 0.6, 0.8, 1])
            Y_uв_uг = np.array([0.35, 0.581, 0.935, 1.170, 1.400, 1.80])
            Z_teta_пл_gamma = np.array([[0.000, 0.032, 0.110, 0.203, 0.259, 0.115],
                                        [0.000, 0.059, 0.157, 0.257, 0.323, 0.193],
                                        [0.000, 0.103, 0.207, 0.310, 0.409, 0.289],
                                        [0.000, 0.103, 0.207, 0.310, 0.409, 0.332],
                                        [0.000, 0.103, 0.207, 0.310, 0.409, 0.381],
                                        [0.000, 0.103, 0.207, 0.310, 0.409, 0.452]])
        if 45 < gamma <= 75:
            X_teta_пл_0 = np.array([0, 0.2, 0.4, 0.6, 0.8, 1])
            Y_uв_uг = np.array([0.35, 0.581, 0.935, 1.170, 1.400, 1.80])
            Z_teta_пл_gamma = np.array([[0.000, 0.000, -0.020, -0.054, -0.071, -0.125],
                                        [0.000, 0.026, 0.093, 0.129, 0.118, 0.079],
                                        [0.000, 0.065, 0.158, 0.252, 0.285, 0.111],
                                        [0.000, 0.065, 0.158, 0.252, 0.337, 0.182],
                                        [0.000, 0.065, 0.158, 0.252, 0.337, 0.277],
                                        [0.000, 0.065, 0.158, 0.252, 0.337, 0.344]])

        if 30 < gamma <= 45:
            X_teta_пл_0 = np.array([0, 0.2, 0.4, 0.6, 0.8, 1])
            Y_uв_uг = np.array([0.35, 0.581, 0.935, 1.170, 1.400, 1.80])
            Z_teta_пл_gamma = np.array([[0.000, 0.031, 0.000, -0.046, 0.016, 0.000],
                                        [0.000, 0.031, 0.000, 0.030, 0.059, 0.017],
                                        [0.000, 0.031, 0.034, 0.097, 0.130, 0.039],
                                        [0.000, 0.031, 0.034, 0.097, 0.153, 0.083],
                                        [0.000, 0.031, 0.034, 0.097, 0.153, 0.083],
                                        [0.000, 0.031, 0.034, 0.097, 0.130, 0.039]])

        F = interpolate.interp2d(X_teta_пл_0, Y_uв_uг, Z_teta_пл_gamma, kind='cubic')(teta_пл_0(x), uв_uг)
    else:
        F = 0 * x
    return F


if __name__ == '__main__':
    ТВД_рез = openpyxl.load_workbook(r'ТВД\Результаты расчета ТВД.xlsx', data_only=True)['Расчет по средней линии тока']

    ТВД_исх = openpyxl.load_workbook(r'ТВД\Исходные данные для расчета ТВД.xlsx', data_only=True)['Турбина']

    ЦИКЛ_исх = openpyxl.load_workbook(r'ЦИКЛ\Исходные данные для цикла.xlsx', data_only=True)['Цикл']

    # Точность итераций
    eps = 10 ** -6

    sigma_КС = ЦИКЛ_исх['B39'].value
    sigma_ОХЛ = 0.98

    # TODO Данные из исходных данных ТВД
    # Допускаемая температура лопатки
    T_лоп_доп = ТВД_исх['B21'].value
    # Температура охлаждающего воздуха в точках выдува
    T_в_torm = ТВД_исх['B6'].value
    # Расход газа на входе в Т
    G_г = ТВД_исх['B5'].value
    # Расход топлива
    G_топ = ТВД_исх['B8'].value
    # Расход охлаждающего воздуха из ТВД
    G_охл_ТВД = ТВД_рез['B38'].value

    # TODO Данные из расчета ТВД
    # Высота лопатки на входе
    h_вх = ТВД_рез['B2'].value * 10 ** -3
    # Высота лопатки на выходе
    h_вых = ТВД_рез['B3'].value * 10 ** -3
    # Температура торможения на входе в турбину
    T_г_torm = ТВД_рез['B26'].value + 200
    # Давление торможения на входе в решетку
    p_г_torm = ТВД_рез['B27'].value * 10 ** 5
    # Скорость газа на входе в решетку
    с_г_вх = ТВД_рез['B30'].value
    # Скорость газа на выходе из решетки
    c_г_вых = ТВД_рез['B35'].value
    # Показатель адиабаты газа на входе в решетку
    k_г = ТВД_рез['B43'].value
    # Плотность газа на выходе из решетки
    ro_г_вых = 1 / ТВД_рез['B53'].value
    # Относительный шаг решетки
    t_отн = ТВД_рез['B57'].value
    # Хорда лопатки
    b = ТВД_рез['B58'].value * 10 ** -3
    # Число лопаток СА
    z_СА = ТВД_рез['B60'].value
    # Приведенная скорость на входе в решетку
    lambda_г_вх = ТВД_рез['B170'].value
    # Приведенная скорость на выходе из решетки
    lambda_вых = ТВД_рез['B171'].value
    # Число Маха на входе в решетку
    M_г_вх = ТВД_рез['B173'].value
    # Число Маха на выходе из решетки
    M_вых = ТВД_рез['B174'].value

    # TODO расчет недостающих данных
    # Газовая постоянная газов
    # Теоретически необходимая масса воздуха для сжигания 1 кг природного газа
    L_0 = 16.683
    # Коэф. избытка воздуха на входе в Т
    alfa = (G_г - G_топ) / (G_топ * L_0)
    R_г = R_продуктов_сгорания_природного_газа(alfa)
    # Статическое давление на входе в решетку
    p_г_вх = p_г_torm * paskal(lambda_г_вх, k_г)
    # Статическая температура на входе в решетку
    T_г_вх = T_г_torm * tau(lambda_г_вх, k_г)
    # Плотность газа на входе в решетку
    ro_г_вх = p_г_вх / (R_г * T_г_вх)
    # Универсальная газовая постоянная
    R_ун = 8.31446261815324
    # Молярная масса воздуха
    mol_возд = 0.02898
    # Газовая постоянная воздуха
    R_в = R_ун / mol_возд
    # Динамическая вязкость газа
    mu_г = mu_г_fun(T_г_torm)
    # Теплопроводность газа
    lambda_г = lambda_г_fun(T_г_torm)
    # Динамическая вязкость охлаждающего воздуха, Па*с
    mu_в = 17.16 * 10 ** -6 * (T_в_torm / 273.15) ** 0.68
    # Теплопроводность охлаждающего воздуха, Вт/(м*К)
    lambda_в = 244.2 * 10 ** -4 * (T_в_torm / 273.15) ** 0.82
    # Показатель адиабаты воздуха
    k_в = Cp_возд(T_в_torm) / (Cp_возд(T_в_torm) - R_в)
    # Давление охлаждающего воздуха
    p_в_torm = p_г_torm / sigma_КС * sigma_ОХЛ

    # TODO Данные решетки №41
    # Максимальная относительная толщина профиля
    с_max_отн = 0.203
    # Угол потока на входе в решетку
    alfa_0 = radians(90)
    # Угол потока на выходе из решетки
    alfa_1 = radians(17.72)
    # Длина спинки
    l_сп = b * 1.217
    # Длина вогнутой поверхности
    l_вп = b * (2.208 - 1.217)
    # Радиус входной кромки
    r_вх = (0.164 * b) / 2
    # Ширина выходного сечения межлопаточного канала
    a_2 = b * 0.185
    # радиус выходной кромки
    r_вых = (a_2 * 0.215) / 2
    # Угол заострения входной кромки
    omega_1 = radians(26.81)
    # Длина дуги входной кромки
    l_вх_кромки = r_вх * (pi - omega_1)
    # Угол установки лопатки
    gamma = radians(44)

    # TODO коэффициенты расходов и потерь
    # Коэф. расхода охлаждающего воздуха #
    mu_расхода_в = 0.82
    # Коэф. потерь давления при истечении из отверстия
    dzeta_охл_D = 0.6

    # TODO Варьируемые исходные данные
    # Степень турбулентности на входной кромке
    Tu_вх = 4
    # Турбулентность на выпуклой и вогнутой поверхностях
    Tu_ср = 2
    # Турбулентность на выходной кромке
    Tu_вых = 5
    # Толщина стенки лопатки, м
    delta_лоп = 1.2 * 10 ** -3
    # Коэф. теплопроводности стенки лопатки при 1173 K, Вт/(м*K)
    lambda_лоп = 24
    # Толщина теплозащитного покрытия
    delta_ТЗП = 0.3 * 10 ** -3
    # Теплопроводность теплозащитного покрытия, Вт/(м*K). В качестве ТЗП выбрана керамика системы Zr-Y-Gd-O TODO
    lambda_ТЗП = 2
    # Относительные координаты точек вдува воздуха
    s_отн_сп = np.array([0.026, 0.22])
    s_отн_вп = np.array([0.032, 0.225, 0.9])
    # Размер пазов в выходной кромке
    h_вых_паза = 2 * 10 ** -3
    delta_вых_паза = 0.55 * 10 ** -3
    t_вых_паза = 1 * 10 ** -3
    # Диаметры отверстий
    d_отв_Л = 0.65 * 10 ** -3
    # Шаг между отверстиями в радиальном направлении
    t_отв_сп = np.array([2.55, 2.6]) * d_отв_Л
    t_отв_вп = np.array([2.55, 2.6]) * d_отв_Л
    gamma_отв_сп = np.array([60, 30])
    gamma_отв_вп = np.array([60, 30, 30])

    # Высота лопатки в рядах отверстий выдува воздуха
    h_лопатки_сп = np.zeros(len(s_отн_сп))
    h_лопатки_вп = np.zeros(len(s_отн_вп))
    # Количество отверстий в ряде
    N_отв_сп = np.zeros(len(s_отн_сп))
    N_отв_вп = np.zeros(len(s_отн_вп))

    for i in range(len(s_отн_сп)):
        h_лопатки_сп[i] = h_вх + (h_вых - h_вх) * s_отн_сп[i]
        N_отв_сп[i] = math.floor(h_лопатки_сп[i] / t_отв_сп[i] + 1)
    for i in range(len(s_отн_вп)):
        h_лопатки_вп[i] = h_вх + (h_вых - h_вх) * s_отн_вп[i]
        if i < len(s_отн_вп) - 1:
            N_отв_вп[i] = math.floor(h_лопатки_вп[i] / t_отв_вп[i] + 1)
        else:
            N_отв_вп[i] = math.floor(h_лопатки_вп[i] / (h_вых_паза + t_вых_паза)) - 1

    # Координаты точек выдува воздуха в абсолютных значениях
    x_s_сп = s_отн_сп * l_сп
    x_s_вп = s_отн_вп * l_вп

    # Критическая скорость звука охлаждающего воздуха
    a_кр_в = sqrt(2 * k_в / (k_в + 1) * R_в * T_в_torm)
    # Критическая скорость звука газов
    a_кр_г = sqrt(2 * k_г / (k_г + 1) * R_г * T_г_torm)

    # Объединяем спинку и корыто, чтобы был единый цикл
    x_s = np.hstack((x_s_сп, x_s_вп))
    s_отн = np.hstack((-s_отн_сп, s_отн_вп))
    N_отв_Л = np.hstack((N_отв_сп, N_отв_вп))
    h_лопатки = np.hstack((h_лопатки_сп, h_лопатки_вп))
    t_отв = np.hstack((t_отв_сп, t_отв_вп))

    gamma_отв = np.hstack((gamma_отв_сп, gamma_отв_вп))

    N_сп = len(s_отн_сп)
    N_вп = len(s_отн_вп)
    N = N_сп + N_вп

    lambda_c_г = np.zeros(N)
    p_г = np.zeros(N)
    T_г = np.zeros(N)
    ro_в = np.zeros(N)
    c_в = np.zeros(N)
    lambda_c_в = np.zeros(N)
    p_в = np.zeros(N)
    T_в = np.zeros(N)
    ro_г = np.zeros(N)
    c_г = np.zeros(N)
    m = np.zeros(N)
    s_э = np.zeros(N)
    Re_s = np.zeros(N)
    A = np.zeros(N)
    teta_пл = np.zeros(N)
    T_пл_torm = np.zeros(N)
    A_1 = np.zeros(N)
    K_в = np.zeros(N)
    G_охл = np.zeros(N)

    delta_teta_пл_gamma = np.zeros(N)
    epsilon_gamma = np.zeros(N)
    teta_пл_gamma = np.zeros(N)

    A_fun = []
    teta_пл_0_fun = []

    delta_teta_пл_gamma_fun = []
    epsilon_gamma_fun = []
    teta_пл_gamma_fun = []
    epsilon_betta_fun = []
    teta_пл_betta_fun = []

    T_пл_torm_fun = []
    A_1_fun = []
    K_в_fun = []

    for i in range(N):
        # Приведенная скорость потока газа в точке выдува
        lambda_c_г[i] = lambda_ад(s_отн[i], lambda_вых)
        # Статическое давление газа в точках выдува воздуха
        p_г[i] = p_г_torm * paskal(lambda_c_г[i], k_г)
        # Статическая температура в точках выдува воздуха
        T_г[i] = T_г_torm * tau(lambda_c_г[i], k_г)
        # Принимается допущение, что полное давление охлаждающего воздуха в канале не изменяется и равно давлению
        # при входе в охлаждающий канал

        cup = 1
        ro_в[i] = 6
        p_в[i] = p_в_torm
        T_в[i] = T_в_torm
        while abs(ro_в[i] - cup) / ro_в[i] * 100 > eps:
            cup = ro_в[i]
            # Плотность охлаждающего воздуха в точке выдува
            ro_в[i] = p_в[i] / (R_в * T_в[i])

            # Скорость истечения охлаждающего воздуха через отверстие
            if x_s[i] <= l_вх_кромки / 2:
                c_в[i] = sqrt(2 * (p_в_torm - p_г_torm) / (ro_в[i] * (1 + dzeta_охл_D)))
                if p_в_torm - p_г_torm <= 0:
                    print_warning(
                        'В точке выдува воздуха с координатами s_отн = ' + str(s_отн[i]) + ' p_г_torm >= p_в_torm')
            else:
                c_в[i] = sqrt(2 * (p_в_torm - p_г[i]) / (ro_в[i] * (1 + dzeta_охл_D)))
                if p_в_torm - p_г[i] <= 0:
                    print_warning(
                        'В точке выдува воздуха с координатами s_отн = ' + str(s_отн[i]) + ' p_г >= p_в_torm')

            if c_в[i] > a_кр_в:
                c_в[i] = a_кр_в

            # Приведенная скорость охлаждающего воздуха через отверстие
            lambda_c_в[i] = c_в[i] / a_кр_в
            # Статическое давление воздуха на выходе из отверстия
            p_в[i] = p_в_torm * paskal(lambda_c_в[i], k_в)
            # Статическая температура охлаждающего воздуха на выходе из отверстия
            T_в[i] = T_в_torm * tau(lambda_c_в[i], k_в)

        # Расход охлаждающего воздуха через ряд отверстий
        if i == N - 1:
            G_охл[i] = mu_расхода_в * N_отв_Л[i] * h_вых_паза * delta_вых_паза * ro_в[i] * c_в[i]
        else:
            G_охл[i] = mu_расхода_в * N_отв_Л[i] * pi * d_отв_Л ** 2 / 4 * ro_в[i] * c_в[i]

    for i in range(N):
        # Плотность газа в точке выдува
        ro_г[i] = p_г[i] / (R_г * T_г[i])
        # Скорость потока газа в точке выдува воздуха
        c_г[i] = lambda_c_г[i] * a_кр_г
        # Параметр выдува воздуха
        m[i] = ro_в[i] * c_в[i] / (ro_г[i] * c_г[i])
        if m[i] < 0.5:
            print('При s_отн = ' + str(s_отн[i]) + ' параметр вдува m = ' + str(m[i]) + ' < 0.5. Происходит отрыв струи.')
        # Эквивалентная высота щели
        if i == N - 1:
            s_э[i] = N_отв_Л[i] * h_вых_паза * delta_вых_паза * 1 / h_лопатки[i]
        else:
            s_э[i] = N_отв_Л[i] * pi * d_отв_Л ** 2 / 4 * 1 / h_лопатки[i]

        # Число Рейнольдса по эквивалентной высоте щели для газа в сечении вдува
        Re_s[i] = c_г[i] * s_э[i] * ro_г[i] / mu_г
        # Температурный фактор
        phi = T_в_torm / T_г_torm

        # Параметр А
        A_fun.append(
            lambda x, k=i: Re_s[k] ** -0.25 * m[k] ** -1.3 * phi ** -1.25 * ((x - x_s[k]) / s_э[k]))

        # Эффективность пленочного охлаждения
        teta_пл_0_fun.append(
            lambda x, k=i: 1 if A_fun[k](x) < 3 else (
                (A_fun[k](x) / 3) ** -0.285 if 3 <= A_fun[k](x) <= 11 else ((A_fun[k](x) / 7.43) ** -0.95)))

        delta_teta_пл_gamma_fun.append(
            lambda x, k=i: delta_teta_пл_gamma_simple_fun(teta_пл_0_fun[k], c_в[k] / c_г[k],
                                                          gamma_отв[k], x))

        epsilon_gamma_fun.append(lambda x, k=i: 1 - delta_teta_пл_gamma_fun[k](x) / teta_пл_0_fun[k](x))

        teta_пл_gamma_fun.append(lambda x, k=i: epsilon_gamma_fun[k](x) * teta_пл_0_fun[k](x))

        epsilon_betta_fun.append(
            lambda x, k=i: 1 - 0.4 if (0 <= x <= x_s[1] and k < N_сп) else 1 - 0.4 if (
                    0 <= x <= x_s[N_сп + 1] and k >= N_сп) else 1 - 0.2)

        teta_пл_betta_fun.append(lambda x, k=i: epsilon_betta_fun[k](x) * teta_пл_gamma_fun[k](x))

        # Температура пленки
        T_пл_torm_fun.append(lambda x, k=i: T_г_torm - teta_пл_betta_fun[k](x) * (T_г_torm - T_в_torm))

        # Влияние вдува на коэф. теплоотдачи #
        A_1_fun.append(lambda x, k=i: ((x - x_s[k]) / s_э[k]) / (c_в[k] / c_г[k]))

        # Влияние выдува воздуха на коэф. теплоотдачи
        K_в_fun.append(lambda x, k=i: 1 + 2 / A_1_fun[k](x))

        # Температура пленки в конце участка
        if i == N_сп - 1:
            A[i] = A_fun[i](l_сп)
            teta_пл[i] = teta_пл_0_fun[i](l_сп)
            T_пл_torm[i] = T_пл_torm_fun[i](l_сп)
            A_1[i] = A_1_fun[i](l_сп)
            K_в[i] = K_в_fun[i](l_сп)
            delta_teta_пл_gamma[i] = delta_teta_пл_gamma_fun[i](l_сп)
            epsilon_gamma[i] = epsilon_gamma_fun[i](l_сп)
            teta_пл_gamma[i] = teta_пл_gamma_fun[i](l_сп)
        elif i == N - 1:
            A[i] = A_fun[i](l_вп)
            teta_пл[i] = teta_пл_0_fun[i](l_вп)
            T_пл_torm[i] = T_пл_torm_fun[i](l_вп)
            A_1[i] = A_1_fun[i](l_вп)
            K_в[i] = K_в_fun[i](l_вп)
            delta_teta_пл_gamma[i] = delta_teta_пл_gamma_fun[i](l_вп)
            epsilon_gamma[i] = epsilon_gamma_fun[i](l_вп)
            teta_пл_gamma[i] = teta_пл_gamma_fun[i](l_вп)
        else:
            A[i] = A_fun[i](x_s[i + 1])
            teta_пл[i] = teta_пл_0_fun[i](x_s[i + 1])
            T_пл_torm[i] = T_пл_torm_fun[i](x_s[i + 1])
            A_1[i] = A_1_fun[i](x_s[i + 1])
            K_в[i] = K_в_fun[i](x_s[i + 1])
            delta_teta_пл_gamma[i] = delta_teta_пл_gamma_fun[i](x_s[i + 1])
            epsilon_gamma[i] = epsilon_gamma_fun[i](x_s[i + 1])
            teta_пл_gamma[i] = teta_пл_gamma_fun[i](x_s[i + 1])

    # Поправка-множитель на степень турбулентности на входной кромке
    if Tu_вх < 1:
        K_Tu_вх = 1
    elif 1 <= Tu_вх <= 4:
        K_Tu_вх = 0.9 * (1 + 0.1 * Tu_вх ** 1.4)
    elif 5 <= Tu_вх <= 10:
        K_Tu_вх = 0.9 * (1 + 0.4 * Tu_вх ** 0.28)

    # Число Рейнольдса на входе в решетку
    Re_вх = с_г_вх * 2 * r_вх * ro_г_вх / mu_г
    # Число Нусельта на входной кромке на непроницаемой поверхности
    Nu_вх_max = Re_вх ** 0.5
    Nu_вх_сред = 0.635 * Re_вх ** 0.5
    # Средний коэф. теплоотдачи газа на входной кромке на непроницаемой поверхности
    alfa_г_вх_max_0 = K_Tu_вх * Nu_вх_max * lambda_г / (2 * r_вх)
    alfa_г_вх_сред_0 = K_Tu_вх * Nu_вх_сред * lambda_г / (2 * r_вх)

    if Re_вх < 5 * 10 ** 3 or Re_вх > 4 * 10 ** 4 or M_г_вх > 0.9 or Tu_вх < 0 or Tu_вх > 10:
        print('Формулы подобия Nu для выходной кромки не справедливы, т.к. нарушаются условия подобия.')

    # Число Рейнольдса осредненное
    Re_ср = c_г_вых * b * ro_г_вых / mu_г
    # Число подобия S_г, предложенное В. И. Локаем (КАИ)
    S_г = sin(alfa_0) / sin(alfa_1) * (
            2 * sin(gamma) / (t_отн * sin(alfa_0 + alfa_1) * cos((alfa_0 - alfa_1) / 2) ** 2) - 1) ** 0.5
    # Число Нусельта для вычисления среднего коэф. теплоотдачи
    Nu_ср = 0.206 * Re_ср ** 0.66 * S_г ** -0.58
    # Поправка-множитель на степень турбулентности на выпуклой и вогнутой поверхностях
    K_Tu_ср = 0.85 * (1 + Tu_ср) ** 0.2
    # Средний по обводу профиля коэф. теплоотдачи от газа к лопатке
    alfa_г_ср_0 = K_Tu_ср * Nu_ср * lambda_г / b

    # Коэф. теплоотдачи от газа к лопатке на вогнутой поверхности 1-1,15
    alfa_г_вп_0 = 1.05 * alfa_г_ср_0
    # Коэф. теплоотдачи от газа к лопатке на спинке на длине от входной кромки до 0.6-0.7 длины спинки 0,75-0,85
    alfa_г_сп_I_0 = 0.8 * alfa_г_ср_0
    # Коэф. теплоотдачи от газа к лопатке на спинке на оставшейся длине 1.2-1,4
    alfa_г_сп_II_0 = 1.2 * alfa_г_ср_0


    # Число Рейнольдса на выходе из решетки
    Re_вых = c_г_вых * b * ro_г_вых / mu_г
    # Число Нусельта на выходной кромке со стороны спинки
    Nu_вых_сп = 0.057 * Re_вых ** 0.7
    # Число Нусельта на выходной кромке со стороны вогнутой поверхности
    Nu_вых_вп = 0.051 * Re_вых ** 0.73

    if Re_вых < 1.86 * 10 ** 5 or Re_вых > 1.5 * 10 ** 6 or M_вых > 0.96 or Tu_вых < 6 or Tu_вых > 12:
        print('Формулы подобия Nu для выходной кромки не справедливы, т.к. нарушаются условия подобия.')

    # Коэф. теплоотдачи газа на выходной кромке со стороны спинки
    alfa_г_вых_сп_0 = Nu_вых_сп * lambda_г / b
    # Коэф. теплоотдачи газа на выходной кромке со стороны вогнутой поверхности
    alfa_г_вых_вп_0 = Nu_вых_вп * lambda_г / b

    # Координаты для построения коэф. теплоотдачи от газа в лопатку
    l_1_сп_г = l_вх_кромки / 2
    l_2_сп_г = l_1_сп_г + 0.6 * l_сп
    l_3_сп_г = l_сп * (1 - 0.15)
    l_4_сп_г = l_сп

    l_1_вп_г = l_вх_кромки / 2
    l_2_вп_г = l_вп * (1 - 0.15)
    l_3_вп_г = l_вп


    def A_single_fun(x):
        return combining_functions(A_fun, x_s, l_сп, l_вп, N_сп, N, A_fun[0](x_s[0]), x)


    def teta_пл_0_single_fun(x):
        return combining_functions(teta_пл_0_fun, x_s, l_сп, l_вп, N_сп, N, teta_пл_0_fun[0](x_s[0]), x)


    def epsilon_gamma_single_fun(x):
        return combining_functions(epsilon_gamma_fun, x_s, l_сп, l_вп, N_сп, N, epsilon_gamma_fun[0](x_s[0]), x)


    def teta_пл_gamma_single_fun(x):
        return combining_functions(teta_пл_gamma_fun, x_s, l_сп, l_вп, N_сп, N, teta_пл_gamma_fun[0](x_s[0]), x)


    def epsilon_betta_single_fun(x):
        return combining_functions(epsilon_betta_fun, x_s, l_сп, l_вп, N_сп, N, epsilon_betta_fun[0](x_s[0]), x)


    def teta_пл_betta_single_fun(x):
        return combining_functions(teta_пл_betta_fun, x_s, l_сп, l_вп, N_сп, N, teta_пл_betta_fun[0](x_s[0]), x)


    # Температура пленки
    def T_пл_torm_single_fun(x):
        return combining_functions(T_пл_torm_fun, x_s, l_сп, l_вп, N_сп, N, T_пл_torm_fun[0](x_s[0]), x)


    def A_1_single_fun(x):
        return combining_functions(A_1_fun, x_s, l_сп, l_вп, N_сп, N, A_1_fun[0](x_s[0]), x)


    # Коэф. увеличения коэф. теплоотдачи при выдуве воздуха из отверстий
    def K_в_single_fun(x):
        return combining_functions(K_в_fun, x_s, l_сп, l_вп, N_сп, N, 1, x)


    # Коэф. теплоотдачи от газа к стенке лопатки без учета выдува
    def alfa_г_0_single_fun(x):
        if -l_4_сп_г <= x < -l_3_сп_г:
            return alfa_г_вых_сп_0
        elif -l_3_сп_г <= x < -l_2_сп_г:
            return alfa_г_сп_II_0
        elif -l_2_сп_г <= x < -l_1_сп_г:
            return alfa_г_сп_I_0
        elif -l_1_сп_г <= x <= -x_s[0]:
            return alfa_г_вх_сред_0
        elif -x_s[0] <= x <= x_s[N_сп]:
            return alfa_г_вх_max_0
        elif x_s[N_сп] <= x < l_1_вп_г:
            return alfa_г_вх_сред_0
        elif l_1_вп_г <= x < l_2_вп_г:
            return alfa_г_вп_0
        elif l_2_вп_г < x < l_3_вп_г:
            return alfa_г_вых_вп_0


    # Коэф. теплоотдачи от газа к стенке лопатки с учетом выдува
    def alfa_г_single_fun(x):
        if -l_сп <= x <= l_вп:
            return alfa_г_0_single_fun(x) * K_в_single_fun(x)


    # Координаты для построения коэф. теплоотдачи от газа в лопатку
    l_1_сп_х = 0.23 * l_сп
    l_2_сп_х = 0.79 * l_сп
    l_3_сп_х = l_сп

    l_1_вп_х = 0.25 * l_вп
    l_2_вп_х = 0.77 * l_вп
    l_3_вп_х = l_вп

    G_охл_сп_1 = 0
    G_охл_вп_1 = 0
    G_охл_сп_2 = 0
    G_охл_вп_2 = 0

    for i in range(N):
        if x_s[i] <= l_1_сп_х and i < len(x_s_сп):
            G_охл_сп_1 += G_охл[i]
        if x_s[i] <= l_1_вп_х and i >= len(x_s_сп):
            G_охл_вп_1 += G_охл[i]
        if l_1_сп_х < x_s[i] <= l_3_сп_х and i < len(x_s_сп):
            G_охл_сп_2 += G_охл[i]
        if l_1_вп_х < x_s[i] <= l_3_вп_х and i >= len(x_s_сп):
            G_охл_вп_2 += G_охл[i]

    G_охл_1 = G_охл_сп_1 + G_охл_вп_1
    G_охл_2 = G_охл_сп_2 + G_охл_вп_2

    # TODO Полость входной кромки дефлектор
    # Диаметр отверстия в дефлекторе
    D_отв_дефлектор_1 = 0.6 * 10 ** -3
    # Шаг отверстий в дефлекторе поперек потока м-у дефлектором и лопаткой
    X_n_1 = D_отв_дефлектор_1 * 2.6
    # Число отверстий в дефлекторе приближенно
    N_отв_Д_1 = math.floor((40 * 10 ** -3 / (X_n_1 * 2)) * (h_лопатки[0] / X_n_1))
    # Приведенная скорость истечения воздуха из отверстия дефлектора через ГДФ расхода охлаждающего воздуха,
    # если он есть
    lambda_x_1 = root(
        lambda lambda_xx: G_охл_1 - m_const(k_в, R_в) * p_в_torm * q(lambda_xx, k_в) \
                          * N_отв_Д_1 * pi * D_отв_дефлектор_1 ** 2 / 4 / sqrt(T_в_torm) * mu_расхода_в,
        np.array(0.5)).x[0]
    # Расстояние от дефлектора до поверхности лопатки
    Z_n_1 = 1 * 10 ** -3
    # Статическое давление воздуха на выходе из отверстия
    p_д_1 = p_в_torm * paskal(lambda_x_1, k_в)
    # Статическая температура охлаждающего воздуха на выходе из отверстия
    T_д_1 = T_в_torm * tau(lambda_x_1, k_в)
    # Плотность охлаждающего воздуха в точке выдува
    ro_в_д_1 = p_д_1 / (R_в * T_д_1)
    # Скорость воздуха через отверстия дефлектора
    W_x_1 = lambda_x_1 * a_кр_в
    # Число Прандля для воздуха
    Pr_возд_1 = mu_в * Cp_возд(T_д_1) / lambda_в
    # Эквивалентный диаметр
    d_э_1 = 2 * Z_n_1
    # Число Рейнольда для воздуха для струйного охлаждения
    Re_D_возд_1 = G_охл_1 * d_э_1 / (h_лопатки[0] * 2 * Z_n_1 * mu_в)

    if Re_D_возд_1 < 3 * 10 ** 2 or Re_D_возд_1 > 3 * 10 ** 4:
        print('Нет данных phi_1 для Re_D_возд_1 = ' + str(Re_D_возд_1) + '. Re_D ∈ [3*10^2..3*10^4]')

    # Комплекс
    V_0_1 = G_охл_1 * R_в * T_в_torm / (p_в_torm * 97 * 10 ** -6)
    Комплекс_1 = W_x_1 / (X_n_1 * 1000 * D_отв_дефлектор_1 * 1000 * ro_в_д_1 * V_0_1)
    # Число Нусельта для струйного охлаждения
    Nu_Dx_1 = phi_1_fun(X_n_1 / D_отв_дефлектор_1, Re_D_возд_1) * phi_2_fun(Комплекс_1, Re_D_возд_1) \
              * Re_D_возд_1 ** m_exp(X_n_1 / D_отв_дефлектор_1, Re_D_возд_1) * Pr_возд_1 ** (1 / 3) \
              * (Z_n_1 / D_отв_дефлектор_1) ** 0.091

    # Nu_Dx_1 = 0.07 * 0.95 * Re_D_возд_1 ** 0.8 * Pr_возд_1 ** (1 / 3) \
    #           * (Z_n_1 / D_отв_дефлектор_1) ** 0.091

    alfa_Dx_1 = Nu_Dx_1 * lambda_в / (2 * Z_n_1)

    # TODO Средняя полость дефлектор
    # Диаметр отверстия в дефлекторе
    D_отв_дефлектор_2 = D_отв_дефлектор_1
    # Шаг отверстий в дефлекторе поперек потока м-у дефлектором и лопаткой
    X_n_2 = X_n_1
    # Число отверстий в дефлекторе приближенно
    N_отв_Д_2 = math.floor((110 * 10 ** -3 / (X_n_2 * 2)) * (h_лопатки[0] / X_n_2))
    # Приведенная скорость истечения воздуха из отверстия дефлектора через ГДФ расхода охлаждающего воздуха,
    # если он есть
    lambda_x_2 = root(
        lambda lambda_xx: G_охл_2 - m_const(k_в, R_в) * p_в_torm * q(lambda_xx, k_в) \
                          * N_отв_Д_2 * pi * D_отв_дефлектор_2 ** 2 / 4 / sqrt(T_в_torm) * mu_расхода_в,
        np.array(0.5)).x[0]
    # Расстояние от дефлектора до поверхности лопатки
    Z_n_2 = 1 * 10 ** -3
    # Статическое давление воздуха на выходе из отверстия
    p_д_2 = p_в_torm * paskal(lambda_x_2, k_в)
    # Статическая температура охлаждающего воздуха на выходе из отверстия
    T_д_2 = T_в_torm * tau(lambda_x_2, k_в)
    # Плотность охлаждающего воздуха в точке выдува
    ro_в_д_2 = p_д_2 / (R_в * T_д_2)
    # Скорость воздуха через отверстия дефлектора
    W_x_2 = lambda_x_2 * a_кр_в
    # Число Прандля для воздуха
    Pr_возд_2 = mu_в * Cp_возд(T_д_2) / lambda_в
    # Эквивалентный диаметр
    d_э_2 = 2 * Z_n_2
    # # Число Рейнольда для воздуха для струйного охлаждения
    Re_D_возд_2 = G_охл_2 * d_э_2 / (h_лопатки[0] * 2 * Z_n_2 * mu_в)

    if Re_D_возд_2 < 3 * 10 ** 2 or Re_D_возд_2 > 3 * 10 ** 4:
        print('Нет данных phi_1 для Re_D_возд_2 = ' + str(Re_D_возд_2) + '. Re_D ∈ [3*10^2..3*10^4]')

    # Комплекс
    V_0_2 = G_охл_2 * R_в * T_в_torm / (p_в_torm * 220 * 10 ** -6)
    Комплекс_2 = W_x_2 / (X_n_2 * 1000 * D_отв_дефлектор_2 * 1000 * ro_в_д_2 * V_0_2)
    # Число Нусельта для струйного охлаждения
    Nu_Dx_2 = phi_1_fun(X_n_2 / D_отв_дефлектор_2, Re_D_возд_2) * phi_2_fun(Комплекс_2, Re_D_возд_2) \
              * Re_D_возд_2 ** m_exp(X_n_2 / D_отв_дефлектор_2, Re_D_возд_2) * Pr_возд_2 ** (1 / 3) \
              * (Z_n_2 / D_отв_дефлектор_2) ** 0.091
    # Nu_Dx_2 = 0.040 * 0.95 * Re_D_возд_2 ** 0.8 * Pr_возд_2 ** (1 / 3) \
    #          * (Z_n_2 / D_отв_дефлектор_2) ** 0.091

    alfa_Dx_2 = Nu_Dx_2 * lambda_в / (2 * Z_n_2)

    #  TODO вых. кромка
    d_экв_паза = 4 * h_вых_паза * delta_вых_паза / (2 * (h_вых_паза + delta_вых_паза))
    Re_вых_паза = c_в[N - 1] * d_экв_паза * ro_в[N - 1] / mu_в
    T_возд_паза = T_в_torm * tau(lambda_c_в[N - 1], k_в)
    Pr_возд_паза = mu_в * Cp_возд(T_возд_паза) / lambda_в
    Nu_вых_х = 0.023 * Re_вых_паза ** 0.8 * Pr_возд_паза ** 0.333
    alfa_вых_х = Nu_вых_х * lambda_в / d_экв_паза


    def delta_лоп_fun(x):
        if -l_3_сп_х <= x < -l_2_сп_х:
            return 3*10**-3
        elif -l_2_сп_х <= x < -l_1_сп_х:
            return delta_лоп
        elif -l_1_сп_х <= x <= -x_s[0]:
            return delta_лоп
        elif -x_s[0] <= x <= x_s[N_сп]:
            return delta_лоп
        elif x_s[N_сп] <= x < l_1_вп_х:
            return delta_лоп
        elif l_1_вп_х <= x < l_2_вп_х:
            return delta_лоп
        elif l_2_вп_х < x < l_3_вп_х:
            return 3*10**-3

    # Коэф. теплоотдачи от воздуха в стенку лопатки
    def alfa_возд_single_fun(x):
        if -l_3_сп_х <= x < -l_2_сп_х:
            return alfa_вых_х
        elif -l_2_сп_х <= x < -l_1_сп_х:
            return alfa_Dx_2
        elif -l_1_сп_х <= x <= -x_s[0]:
            return alfa_Dx_1
        elif -x_s[0] <= x <= x_s[N_сп]:
            return alfa_Dx_1
        elif x_s[N_сп] <= x < l_1_вп_х:
            return alfa_Dx_1
        elif l_1_вп_х <= x < l_2_вп_х:
            return alfa_Dx_2
        elif l_2_вп_х < x < l_3_вп_х:
            return alfa_вых_х


    # Коэф. теплопередачи
    def K_single_fun(x):
        if -l_сп <= x <= l_вп:
            return (1 / alfa_г_single_fun(
                x) + delta_лоп / lambda_лоп + delta_ТЗП / lambda_ТЗП + 1 / alfa_возд_single_fun(x)) ** -1


    # Тепловой поток
    def Q(x):
        if -l_сп <= x <= l_вп:
            return (T_пл_torm_single_fun(x) - T_в_torm) * K_single_fun(x)


    # Коэф. теплопередачи от газа в стенку лопатки
    def K_ст_г_single_fun(x):
        if -l_сп <= x <= l_вп:
            return (delta_ТЗП / lambda_ТЗП + 1 / alfa_г_single_fun(x)) ** -1


    # Коэф. теплопередачи от воздуха в стенку лопатки
    def K_ст_охл_single_fun(x):
        if -l_сп <= x <= l_вп:
            return (1 / alfa_возд_single_fun(x)) ** -1


    # Температура лопатки со стороны газа
    def T_лоп_ст_г_single_fun(x):
        if -l_сп <= x <= l_вп:
            return T_пл_torm_single_fun(x) - Q(x) / K_ст_г_single_fun(x)


    # Температура лопатки со стороны воздуха
    def T_лоп_ст_охл_single_fun(x):
        if -l_сп <= x <= l_вп:
            return T_в_torm + Q(x) / K_ст_охл_single_fun(x)


    # Дискретность построения графика
    dis = 0.1 * 10 ** -3
    font_size = 17
    X = np.arange(-l_сп, l_вп + dis, dis)

    Y = np.zeros(len(X))
    Y2 = np.zeros(len(X))
    Y3 = np.zeros(len(X))
    for i in range(len(X)):
        Y[i] = T_пл_torm_single_fun(X[i])
        Y2[i] = T_лоп_ст_г_single_fun(X[i])
        Y3[i] = T_лоп_ст_охл_single_fun(X[i])
    fig1, ax1 = plt.subplots(1, figsize=(400 / 25.4, 220 / 25.4))
    l1, = ax1.plot(X * 1000, Y, color='black', linestyle=':')
    l2, = ax1.plot(X * 1000, Y2, color='black')
    l3, = ax1.plot(X * 1000, Y3, color='black', linestyle='--')
    l4, = ax1.plot([min(X * 1000), max(X * 1000)], [T_лоп_доп, T_лоп_доп], color='black',
                   linestyle='-.', linewidth=2)
    ax1.grid()
    ax1.minorticks_on()
    ax1.grid(which='major', linewidth=0.7, color='k')
    ax1.grid(which='minor', linewidth=0.5, linestyle='--')
    min_x = floor(min(X * 1000), 10)
    max_x = ceil(max(X * 1000), 10)
    ax1.set_xlim(min_x, max_x)
    ax1.set_xticks(np.arange(min_x, max_x + 10, 10))
    min_y = floor(min([min(Y), min(Y2)]), 100)
    max_y = ceil(max([max(Y), max(Y2)]), 100)
    ax1.set_ylim(min_y, max_y)
    ax1.set_yticks(np.arange(min_y, max_y + 100, 100))
    ax1.legend([l1, l2, l3, l4], [r'$T_{пл}^*$', r'$T_{ст_г}$', r'$T_{ст_х}$', r'$T_{л_{доп}}$'], ncols=4, loc=9,
               fontsize=font_size)
    ax1.tick_params(axis='both', labelsize=font_size)
    ax1.text(min_x + 5, min_y + 10, r'$l_{сп}$', fontsize=font_size)
    ax1.text(max_x - 5, min_y + 10, r'$l_{к}$', fontsize=font_size)
    ax1.set_xlabel(r'$l, мм$', fontsize=font_size)
    ax1.set_ylabel(r'$T, K$', fontsize=font_size)
    fig1.savefig(r'Охлаждение\Температура пленки и лопатки.jpg', dpi=800)

    Y = np.zeros(len(X))
    Y2 = np.zeros(len(X))
    for i in range(len(X)):
        Y[i] = alfa_г_0_single_fun(X[i])
        Y2[i] = alfa_возд_single_fun(X[i])
    fig1, ax1 = plt.subplots(1, figsize=(400 / 25.4, 220 / 25.4))
    fig1.subplots_adjust(hspace=10 / 25.4)
    l1, = ax1.plot(X * 1000, Y, color='black')
    l2, = ax1.plot(X * 1000, Y2, color='black', linestyle='--')
    ax1.grid()
    ax1.minorticks_on()
    ax1.grid(which='major', linewidth=0.7, color='k')
    ax1.grid(which='minor', linewidth=0.5, linestyle='--')
    min_x = floor(min(X * 1000), 10)
    max_x = ceil(max(X * 1000), 10)
    ax1.set_xlim(min_x, max_x)
    ax1.set_xticks(np.arange(min_x, max_x + 10, 10))
    min_y = floor(min([min(Y), min(Y2)]), 1000)
    max_y = ceil(max([max(Y), max(Y2)]), 1000)
    ax1.set_ylim(min_y, max_y)
    ax1.set_yticks(np.arange(min_y, max_y + 50, 1000))
    ax1.legend([l1, l2], [r'$\alpha_{г0}^*$', r'$\alpha_{в}$'], loc=9,
               fontsize=font_size)
    ax1.tick_params(axis='both', labelsize=font_size)
    ax1.text(min_x + 5, min_y + 200, r'$l_{сп}$', fontsize=font_size)
    ax1.text(max_x - 5, min_y + 200, r'$l_{к}$', fontsize=font_size)
    ax1.set_xlabel(r'$l, мм$', fontsize=font_size)
    ax1.set_ylabel(r'$\alpha, \frac{Вт}{м^2*К}$', fontsize=font_size)
    fig1.savefig(r'Охлаждение\Коэф. теплоотдачи газа без учета выдува воздуха.jpg', dpi=800)

    plt.close('all')

    s_отн = my_flip_matr(s_отн, N_сп)
    x_s = my_flip_matr(np.hstack((-x_s_сп, x_s_вп)), N_сп)
    N_отв_Л = my_flip_matr(N_отв_Л, N_сп)
    h_лопатки = my_flip_matr(h_лопатки, N_сп)
    lambda_c_г = my_flip_matr(lambda_c_г, N_сп)
    p_г = my_flip_matr(p_г, N_сп)
    T_г = my_flip_matr(T_г, N_сп)
    ro_в = my_flip_matr(ro_в, N_сп)
    c_в = my_flip_matr(c_в, N_сп)
    lambda_c_в = my_flip_matr(lambda_c_в, N_сп)
    p_в = my_flip_matr(p_в, N_сп)
    T_в = my_flip_matr(T_в, N_сп)
    ro_г = my_flip_matr(ro_г, N_сп)
    c_г = my_flip_matr(c_г, N_сп)
    m = my_flip_matr(m, N_сп)
    s_э = my_flip_matr(s_э, N_сп)
    Re_s = my_flip_matr(Re_s, N_сп)
    A = my_flip_matr(A, N_сп)
    teta_пл = my_flip_matr(teta_пл, N_сп)
    T_пл_torm = my_flip_matr(T_пл_torm, N_сп)
    G_охл = my_flip_matr(G_охл, N_сп)
    A_1 = my_flip_matr(A_1, N_сп)
    K_в = my_flip_matr(K_в, N_сп)

    delta_teta_пл_gamma = my_flip_matr(delta_teta_пл_gamma, N_сп)
    epsilon_gamma = my_flip_matr(epsilon_gamma, N_сп)
    teta_пл_gamma = my_flip_matr(teta_пл_gamma, N_сп)

    Excel_ОХЛАЖДЕНИЕ = pd.DataFrame({
        's_отн': s_отн,
        'x_s, мм': x_s * 1000,
        'N_отв': N_отв_Л,
        'h_лопатки, мм': h_лопатки * 1000,
        'lambda_c_г': lambda_c_г,
        'p_г, Па*10^5': p_г * 10 ** -5,
        'T_г, К': T_г,
        'ro_в, кг/м^3': ro_в,
        'c_в, м/с': c_в,
        'lambda_c_в': lambda_c_в,
        'p_в, Па*10^5': p_в * 10 ** -5,
        'T_в, К': T_в,
        'ro_г, кг/м^3': ro_г,
        'c_г, м/с': c_г,
        'm': m,
        's_э, мм': s_э * 1000,
        'Re_s': Re_s,
        'A': A,
        'teta_пл': teta_пл,
        'T_пл_torm, К': T_пл_torm,
        'G, кг/с': G_охл * 1000,
        'g, %': G_охл / G_г * 100,
        'A_1': A_1,
        'K_в': K_в,
        'delta_teta_пл_gamma': delta_teta_пл_gamma,
        'epsilon_gamma': epsilon_gamma,
        'teta_пл_gamma': teta_пл_gamma
    }).round(3).transpose()

    # TODO Точки профиля
    x_отн = np.array(
        [0.00, 0.01, 0.02, 0.03, 0.05, 0.07, 0.09, 0.12, 0.15, 0.18, 0.21, 0.24, 0.27, 0.30,
         0.34, 0.38, 0.42, 0.46, 0.50, 0.54, 0.58, 0.62, 0.66, 0.70, 0.74, 0.78, 0.81, 0.84,
         0.87, 0.90, 0.92, 0.94, 0.96, 0.98, 1.00])
    y_к_отн = np.array(
        [0.0834, 0.0573, 0.0512, 0.0452, 0.0331, 0.0209, 0.0086, 0.0094, 0.0274, 0.0439, 0.0588, 0.0713, 0.0810, 0.0892,
         0.0977, 0.1035, 0.1069, 0.1078, 0.1066, 0.1034, 0.0985, 0.0922, 0.0848, 0.0765, 0.0674, 0.0575, 0.0496, 0.0413,
         0.0327, 0.0241, 0.0182, 0.0122, 0.0062, 0.0003, -0.0057])
    y_сп_отн = np.array([
        0.1044, 0.1212, 0.1378, 0.1533, 0.1795, 0.2015, 0.2190, 0.2390, 0.2542, 0.2651, 0.2724, 0.2768, 0.2787, 0.2786,
        0.2757, 0.2703, 0.2627, 0.2532, 0.2420, 0.2294, 0.2158, 0.2018, 0.1869, 0.1719, 0.1563, 0.1402, 0.1276, 0.1143,
        0.1003, 0.0854, 0.0753, 0.0650, 0.0547, 0.0439, 0.0324])

    [X_к, Y_к] = rotation_matrix(-(gamma + pi), x_отн * b, 0, y_к_отн * b, 0)
    [X_сп, Y_сп] = rotation_matrix(-(gamma + pi), x_отн * b, 0, y_сп_отн * b, 0)

    Excel_Точки_корыта = r'Охлаждение\Точки корыта.csv'
    Excel_Точки_спинки = r'Охлаждение\Точки спинки.csv'
    Excel_Точки_корыта_1 = r'Охлаждение\Точки корыта.xlsx'
    Excel_Точки_спинки_1 = r'Охлаждение\Точки спинки.xlsx'

    Точки_корыта = pd.DataFrame({
        'X_к': X_к[7:len(X_к) - 1] * 1000,
        'Y_к': Y_к[7:len(X_к) - 1] * 1000,
    }).round(3)

    Точки_спинки = pd.DataFrame({
        'X_сп': X_сп * 1000,
        'Y_сп': Y_сп * 1000,
    }).round(3)

    pd.options.display.width = None
    pd.options.display.max_columns = None
    # Сброс ограничений на количество выводимых рядов
    pd.set_option('display.max_rows', None)
    # Сброс ограничений на число столбцов
    pd.set_option('display.max_columns', None)

    print(' ')
    print('****************************************************************')
    print(Excel_ОХЛАЖДЕНИЕ)
    print('****************************************************************')
    print(' ')

    if os.path.exists(Excel_Точки_корыта):
        os.remove(Excel_Точки_корыта)
    Точки_корыта.to_csv(Excel_Точки_корыта, header=False, index=False)

    if os.path.exists(Excel_Точки_спинки):
        os.remove(Excel_Точки_спинки)
    Точки_спинки.to_csv(Excel_Точки_спинки, header=False, index=False)

    if os.path.exists(Excel_Точки_корыта_1):
        os.remove(Excel_Точки_корыта_1)
    Точки_корыта.to_excel(Excel_Точки_корыта_1, header=False, index=False)

    if os.path.exists(Excel_Точки_спинки_1):
        os.remove(Excel_Точки_спинки_1)
    Точки_спинки.to_excel(Excel_Точки_спинки_1, header=False, index=False)

    os.startfile(r'Охлаждение\Температура пленки и лопатки.jpg')
    os.startfile(r'Охлаждение\Коэф. теплоотдачи газа без учета выдува воздуха.jpg')
    # os.startfile(r'Охлаждение\Коэф. теплоотдачи газа с учетом выдува воздуха.jpg')

    print('Расход через 1 полость', G_охл_1, ' кг/с')
    print('Расход через 2 полость', G_охл_2, ' кг/с')
    print('Расход через 1 СА = ', sum(G_охл), ' кг/с')
    print('Суммарный расход через все СА = ', sum(G_охл) * z_СА, ' кг/с')
    print('Суммарный расход через все СА из расч. ТВД = ', G_охл_ТВД, ' кг/с')
